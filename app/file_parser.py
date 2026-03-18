"""
리뷰 데이터 파서 (xlsx, csv)
"""
import csv
import io
from pathlib import Path


def extract_reviews(file_path: str) -> str:
    """xlsx 또는 csv 파일에서 리뷰 데이터를 추출하여 마크다운으로 반환"""
    path = Path(file_path)
    suffix = path.suffix.lower()

    try:
        if suffix == ".xlsx":
            return _parse_xlsx(file_path)
        elif suffix == ".csv":
            return _parse_csv(file_path)
        else:
            return ""
    except Exception as e:
        print(f"[WARN] Review parsing failed: {e}")
        return ""


def _parse_xlsx(file_path: str) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(file_path, read_only=True)
    ws = wb.active

    rows = []
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c) if c else f"col_{j}" for j, c in enumerate(row)]
            continue
        if i > 500:  # 최대 500개 리뷰
            break
        rows.append(row)

    wb.close()
    return _format_reviews(headers, rows)


def _parse_csv(file_path: str) -> str:
    with open(file_path, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        headers = next(reader, None)
        if not headers:
            return ""
        rows = []
        for i, row in enumerate(reader):
            if i >= 500:
                break
            rows.append(row)

    return _format_reviews(headers, rows)


def _format_reviews(headers: list, rows: list) -> str:
    if not headers or not rows:
        return ""

    lines = [
        f"# 리뷰 데이터 ({len(rows)}건)\n",
        f"## 컬럼: {', '.join(headers)}\n",
    ]

    # 리뷰 요약 통계
    lines.append(f"총 리뷰 수: {len(rows)}건\n")

    # 별점 컬럼 자동 감지 (rating, 별점, 평점, score 등)
    rating_idx = None
    for i, h in enumerate(headers):
        if any(k in h.lower() for k in ["rating", "별점", "평점", "score", "점수", "star"]):
            rating_idx = i
            break

    if rating_idx is not None:
        ratings = []
        for row in rows:
            try:
                r = float(row[rating_idx])
                ratings.append(r)
            except (ValueError, TypeError, IndexError):
                pass
        if ratings:
            avg = sum(ratings) / len(ratings)
            lines.append(f"평균 별점: {avg:.1f} / 5.0")
            lines.append(f"5점: {sum(1 for r in ratings if r >= 4.5)}건, "
                         f"4점: {sum(1 for r in ratings if 3.5 <= r < 4.5)}건, "
                         f"3점 이하: {sum(1 for r in ratings if r < 3.5)}건\n")

    # 리뷰 텍스트 컬럼 자동 감지
    text_idx = None
    for i, h in enumerate(headers):
        if any(k in h.lower() for k in ["review", "리뷰", "내용", "comment", "text", "본문", "후기"]):
            text_idx = i
            break

    # 상위 리뷰 샘플 (최대 30개)
    lines.append("\n## 리뷰 샘플\n")
    sample_count = min(30, len(rows))
    for i, row in enumerate(rows[:sample_count]):
        parts = []
        if rating_idx is not None and rating_idx < len(row):
            parts.append(f"★{row[rating_idx]}")
        if text_idx is not None and text_idx < len(row):
            text = str(row[text_idx])[:200]
            parts.append(text)
        else:
            # 모든 컬럼 결합
            parts.append(" | ".join(str(c)[:100] for c in row if c))
        lines.append(f"{i+1}. {' — '.join(parts)}")

    return "\n".join(lines)
